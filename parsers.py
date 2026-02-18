#!/usr/bin/env python3
"""
Excel file parsing functions for the dashboard generator.
"""

import pandas as pd
import openpyxl
import sys
from datetime import datetime


def parse_amount(value):
    """Parse amount values that may contain currency symbols and formatting."""
    if pd.isna(value) or value == '':
        return 0.0

    # Convert to string and clean
    if isinstance(value, (int, float)):
        return float(value)

    value_str = str(value).strip()

    # Remove currency symbols and spaces
    value_str = value_str.replace('€', '').replace('$', '').replace(',', '').strip()

    try:
        return float(value_str)
    except ValueError:
        return 0.0


def parse_hours(value):
    """Parse hours values."""
    if pd.isna(value) or value == '':
        return 0.0

    if isinstance(value, (int, float)):
        return float(value)

    try:
        return float(str(value).strip())
    except ValueError:
        return 0.0


def parse_rate(value):
    """Parse rate values (hourly rate or percentage)."""
    if pd.isna(value) or value == '':
        return None

    value_str = str(value).strip()

    # Remove currency symbols
    value_str = value_str.replace('€', '').replace('$', '').replace(',', '').strip()

    # Check if it's a percentage
    if '%' in value_str:
        try:
            return float(value_str.replace('%', ''))
        except ValueError:
            return None

    try:
        return float(value_str)
    except ValueError:
        return None


def _find_column_index(header_row, name):
    """Find column index by header name (case-insensitive, stripped)."""
    for i, cell in enumerate(header_row):
        if cell.value and str(cell.value).strip().lower() == name.lower():
            return i
    return None


def _extract_comments(ws, col_map, header_row_idx):
    """Extract all cell comments from a worksheet, keyed by (row, col_name)."""
    comments = {}
    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        row_comments = {}
        for col_name, col_idx in col_map.items():
            cell = ws.cell(row=row_idx, column=col_idx + 1)  # openpyxl is 1-indexed
            if cell.comment:
                row_comments[col_name] = cell.comment.text.strip()
        if row_comments:
            comments[row_idx] = row_comments
    return comments


def read_excel_data(file_path):
    """Read and parse Excel file, reading all sheets."""
    try:
        # Use openpyxl to get sheet names and comments
        wb = openpyxl.load_workbook(file_path)
        sheet_names = wb.sheetnames

        all_records = []
        global_index = 1

        for sheet_name in sheet_names:
            ws = wb[sheet_name]

            # Find header row (first row with expected column names)
            header_row_idx = 0
            header_row = list(ws.iter_rows(min_row=1, max_row=1))[0]

            expected_cols = ['Date', 'Event Type', 'Project', 'Hourly Rate', 'Additional Rate', 'Hours', 'Amount']
            col_map = {}
            for name in ['#'] + expected_cols + ['Comment']:
                idx = _find_column_index(header_row, name)
                if idx is not None:
                    col_map[name] = idx

            missing_cols = [c for c in expected_cols if c not in col_map]
            if missing_cols:
                print(f"Warning: Sheet '{sheet_name}' missing columns: {missing_cols}")
                available = [str(c.value).strip() for c in header_row if c.value]
                print(f"  Available columns: {available}")

            # Extract cell comments from this sheet
            cell_comments = _extract_comments(ws, col_map, 1)

            # Read data with pandas for the actual parsing
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            df.columns = df.columns.str.strip()

            for idx, row in df.iterrows():
                # Skip empty rows
                if pd.isna(row.get('Project', '')) or str(row.get('Project', '')).strip() == '':
                    continue

                billing_rate = parse_rate(row.get('Hourly Rate'))
                surcharge_rate = parse_rate(row.get('Additional Rate'))
                hours = parse_hours(row.get('Hours'))
                amount = parse_amount(row.get('Amount'))
                event_type = str(row.get('Event Type', '')).strip() if pd.notna(row.get('Event Type')) else ''

                # Calculate billable amount for Working Time: hours × billing_rate × (1 + surcharge_rate)
                billable_amount = None
                if event_type == 'Working Time' and billing_rate is not None and hours > 0:
                    multiplier = 1.0
                    if surcharge_rate is not None:
                        multiplier = 1.0 + surcharge_rate
                    billable_amount = hours * billing_rate * multiplier

                # Build comment: combine explicit Comment column + cell comments
                comment_parts = []
                # Explicit Comment column value
                comment_val = row.get('Comment')
                if pd.notna(comment_val) and str(comment_val).strip():
                    comment_parts.append(str(comment_val).strip())
                # Cell-level comments from openpyxl (row in Excel is idx+2: +1 for 0-index, +1 for header)
                excel_row = idx + 2
                if excel_row in cell_comments:
                    for col_name, comment_text in cell_comments[excel_row].items():
                        comment_parts.append(f"{col_name}: {comment_text}")

                comment = ' | '.join(comment_parts) if comment_parts else ''

                record = {
                    'index': global_index,
                    'date': str(row.get('Date', '')) if pd.notna(row.get('Date')) else '',
                    'event_type': event_type,
                    'project': str(row.get('Project', '')).strip() if pd.notna(row.get('Project')) else '',
                    'billing_rate': billing_rate,
                    'surcharge_rate': surcharge_rate,
                    'hours': hours,
                    'amount': amount,
                    'billable_amount': billable_amount,
                    'comment': comment,
                    'sheet': sheet_name
                }
                global_index += 1

                # Parse date
                if record['date']:
                    try:
                        if hasattr(record['date'], 'strftime'):
                            record['date'] = record['date'].strftime('%Y-%m-%d')
                        elif isinstance(record['date'], datetime):
                            record['date'] = record['date'].strftime('%Y-%m-%d')
                        elif isinstance(record['date'], str):
                            date_str = record['date'].strip()
                            if ' ' in date_str:
                                date_str = date_str.split(' ')[0]
                            for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%Y/%m/%d']:
                                try:
                                    dt = datetime.strptime(date_str, fmt)
                                    record['date'] = dt.strftime('%Y-%m-%d')
                                    break
                                except ValueError:
                                    continue
                            if len(date_str) == 10 and date_str[4] == '-' and date_str[7] == '-':
                                try:
                                    datetime.strptime(date_str, '%Y-%m-%d')
                                    record['date'] = date_str
                                except ValueError:
                                    pass
                    except Exception:
                        pass

                # Extract year-month for monthly summaries
                if record['date']:
                    try:
                        record['year_month'] = record['date'][:7]
                    except:
                        record['year_month'] = ''
                else:
                    record['year_month'] = ''

                all_records.append(record)

        wb.close()
        return all_records

    except Exception as e:
        print(f"Error reading Excel file: {e}")
        sys.exit(1)
